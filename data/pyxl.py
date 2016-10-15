# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook
# wb = load_workbook(filename='fullTejNav.xlsx', read_only=True)
wb = load_workbook(filename='b2bcnyes.xlsx', read_only=True)
ws = wb['Sheet1'] # ws is now an IterableWorksheet

for row in ws.rows:
    for i in xrange(len(row)):
        if isinstance(row[i].value, str) or isinstance(row[i].value, unicode):
            value = str(row[i].value.encode('utf-8'))
        elif row[i].value is None:
            value = ''
        else:
            value = str(row[i].value)

        sys.stdout.write(value)
        if (i < len(row) - 1):
            sys.stdout.write(',')
    print ''
